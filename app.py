import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_exel(file_name, sheet_name):
    wb = xl.load_workbook(file_name)
    sheet = wb[sheet_name]
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "e2")

    wb.save(f"updated-{file_name}")


try:
    file = input("enter the name of the exel file> ")
    sheet = input("enter the name of the exel sheet> ")
    print("working on the file")
    process_exel(file, sheet)
    print("Done")
except:
    print("wrong")
